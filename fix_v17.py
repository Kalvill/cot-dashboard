#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ДІАГНОСТИКА v21 — три проблеми:
  1. NZD: COT INDEX LS = 0.0 в overview
  2. NZD: TFF Report неактивний (вкладка не знайдена)
  3. GOLD: Disaggregated показує нулі
Запусти цей скрипт і надішли весь вивід з консолі.
"""
import pandas as pd
from pathlib import Path

BASE = Path(__file__).parent
OVERVIEW_FILE = BASE / "data" / "COT_OVERVIEW.xlsx"
TFF_FILE      = BASE / "data" / "COT_TFF_REPORTS.xlsx"
DISAG_FILE    = BASE / "data" / "COT_DISAGRAGATE_REPORTS.xlsx"

print("=" * 62)
print("  ДІАГНОСТИКА v21")
print("=" * 62)

# ────────────────────────────────────────────────────────────
# 1. NZD в overview — що лежить у колонках 2..14
# ────────────────────────────────────────────────────────────
print("\n█ 1. NZD в COT_OVERVIEW.xlsx (вкладка overview)")
if not OVERVIEW_FILE.exists():
    print("  ✗ Файл не знайдено!")
else:
    xl = pd.ExcelFile(OVERVIEW_FILE)
    raw = xl.parse('overview', header=None)
    found = False
    for i in range(4, len(raw)):
        asset = str(raw.iloc[i, 1]).strip()
        if asset.upper() == 'NZD':
            found = True
            print(f"  Рядок {i} (Excel рядок {i+1}):")
            for c in range(0, 20):
                if c < raw.shape[1]:
                    print(f"    кол {c:2d}: {repr(raw.iloc[i, c])}")
    if not found:
        print("  ✗ Рядок NZD не знайдено у колонці B!")

    # 1б. NZD вкладка — останні значення LS NET, щоб перевірити чи 0.0 легітимний
    if 'NZD' in xl.sheet_names:
        print("\n█ 1б. Вкладка NZD — LS NET за останні 5 тижнів (колонка 8)")
        raw2 = xl.parse('NZD', header=None)
        df = raw2.iloc[5:].copy()
        dates = pd.to_datetime(df.iloc[:, 1], errors='coerce')
        ok = dates.notna() & (dates.dt.year > 2000)
        df = df[ok]
        ls_net = pd.to_numeric(df.iloc[:, 8], errors='coerce')
        mn, mx = ls_net.min(), ls_net.max()
        cur = ls_net.iloc[-1]
        print(f"    Останні 5: {ls_net.tail(5).tolist()}")
        print(f"    MIN={mn}  MAX={mx}  CUR={cur}")
        if mx != mn:
            idx = (cur - mn) / (mx - mn) * 100
            print(f"    → COT Index розрахунково = {idx:.1f}%")
            print(f"    (якщо ≈0 — значення легітимне: LS на історичному мінімумі)")

# ────────────────────────────────────────────────────────────
# 2. TFF — список вкладок і чи є щось схоже на NZD
# ────────────────────────────────────────────────────────────
print("\n█ 2. Вкладки у COT_TFF_REPORTS.xlsx")
if not TFF_FILE.exists():
    print("  ✗ Файл не знайдено!")
else:
    xl_t = pd.ExcelFile(TFF_FILE)
    for s in xl_t.sheet_names:
        mark = "  ← схоже на NZD?" if 'NZ' in s.upper() or 'KIWI' in s.upper() else ""
        print(f"    {repr(s)}{mark}")

# ────────────────────────────────────────────────────────────
# 3. DISAG — GOLD: сирі значення в колонках MM/PM/SD
# ────────────────────────────────────────────────────────────
print("\n█ 3. COT_DISAGRAGATE_REPORTS.xlsx — вкладки та GOLD")
if not DISAG_FILE.exists():
    print("  ✗ Файл не знайдено!")
else:
    xl_d = pd.ExcelFile(DISAG_FILE)
    print("  Вкладки:")
    for s in xl_d.sheet_names:
        print(f"    {repr(s)}")

    gold_sheet = None
    for s in xl_d.sheet_names:
        if s.strip().upper() == 'GOLD':
            gold_sheet = s
            break
    if gold_sheet is None:
        print("\n  ✗ Вкладка GOLD не знайдена!")
    else:
        print(f"\n  Вкладка {repr(gold_sheet)} — сирі клітинки:")
        raw_g = xl_d.parse(gold_sheet, header=None)
        print(f"    Розмір: {raw_g.shape[0]} рядків × {raw_g.shape[1]} колонок")
        # Показуємо рядки 18-25 (навколо DISAG_DATA_START=20), колонки date/mm/pm/sd/oi
        cols_check = {'date': 1, 'mm_cl': 4, 'mm_cs': 5, 'mm_net': 8,
                      'pm_net': 15, 'sd_net': 22, 'oi': 34}
        for ri in range(15, min(28, len(raw_g))):
            vals = []
            for name, ci in cols_check.items():
                v = raw_g.iloc[ri, ci] if ci < raw_g.shape[1] else '—'
                vals.append(f"{name}={repr(v)}")
            print(f"    рядок {ri}: " + " | ".join(vals))

        # Перевірка через openpyxl: формули чи значення?
        print("\n  Перевірка формул через openpyxl (data_only=False):")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(DISAG_FILE, data_only=False, read_only=True)
            ws = wb[gold_sheet]
            # Excel рядок = pandas індекс + 1
            for excel_row in (21, 22, 23):
                cell_e = ws.cell(row=excel_row, column=5)   # кол E = mm_cl (індекс 4)
                cell_i = ws.cell(row=excel_row, column=9)   # кол I = mm_net (індекс 8)
                print(f"    E{excel_row}: {repr(cell_e.value)}  |  I{excel_row}: {repr(cell_i.value)}")
            wb.close()
            wb2 = load_workbook(DISAG_FILE, data_only=True, read_only=True)
            ws2 = wb2[gold_sheet]
            print("  Те саме з data_only=True (кешовані значення):")
            for excel_row in (21, 22, 23):
                cell_e = ws2.cell(row=excel_row, column=5)
                cell_i = ws2.cell(row=excel_row, column=9)
                print(f"    E{excel_row}: {repr(cell_e.value)}  |  I{excel_row}: {repr(cell_i.value)}")
            wb2.close()
        except Exception as e:
            print(f"    ⚠ openpyxl: {e}")

print("\n" + "=" * 62)
print("  Готово. Надішли весь вивід.")
print("=" * 62)